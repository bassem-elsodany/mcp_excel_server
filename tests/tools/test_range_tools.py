"""
Test cases for Excel MCP Server range tools.

This module contains test cases for all range tools including:
- delete_range_tool
- copy_range_tool
- move_range_tool
- merge_range_tool
- unmerge_range_tool
"""

import os
import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment
from mcp_excel_server.config.settings import settings
from mcp.server.fastmcp import FastMCP
from mcp_excel_server.core.exceptions import ValidationError, SheetError, RangeError

from mcp_excel_server.api.tools.range import (
    delete_range_tool,
    copy_range_tool,
    move_range_tool,
    merge_range_tool,
    unmerge_range_tool
)

# Test data
TEST_WORKBOOK = f"{settings.excel_mcp_folder}/test_workbook.xlsx"
TEST_INVALID_WORKBOOK = f"{settings.excel_mcp_folder}/invalid/name.xlsx"
TEST_NONE_EXISTENT_WORKBOOK = f"{settings.excel_mcp_folder}/nonexistent.xlsx"
TEST_SHEET = "Sheet1"
TEST_RANGE = "A1:B2"
TEST_CELL = "A1"
TEST_TARGET = "C3"

@pytest.fixture
def mock_mcp():
    """Create a mock FastMCP server instance."""
    return Mock(spec=FastMCP)

@pytest.fixture
def test_workbook():
    """Create a test workbook with some data."""
    # Ensure test directory exists
    os.makedirs(settings.excel_mcp_folder, exist_ok=True)
    
    # Create test workbook
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = TEST_SHEET
    
    # Add some test data
    ws['A1'] = "Test1"
    ws['A2'] = "Test2"
    ws['B1'] = "Test3"
    ws['B2'] = "Test4"
    
    # Save workbook
    wb.save(TEST_WORKBOOK)
    
    yield TEST_WORKBOOK
    
    # Cleanup
    if os.path.exists(TEST_WORKBOOK):
        os.remove(TEST_WORKBOOK)

class TestDeleteRangeTool:
    """Test cases for delete_range_tool."""
    
    def test_delete_range_success(self, mock_mcp, test_workbook):
        """Test successful range deletion."""
        result = delete_range_tool(TEST_WORKBOOK, TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Deleted range" in result["message"]
        
    def test_delete_range_invalid_sheet(self, mock_mcp, test_workbook):
        """Test deleting range from invalid sheet."""
        result = delete_range_tool(TEST_WORKBOOK, "InvalidSheet", "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "Sheet" in result["message"]
        
    def test_delete_range_invalid_range(self, mock_mcp, test_workbook):
        """Test deleting invalid range."""
        result = delete_range_tool(TEST_WORKBOOK, TEST_SHEET, "InvalidRange", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "error" in result["message"].lower()

class TestCopyRangeTool:
    """Test cases for copy_range_tool."""
    
    def test_copy_range_success(self, mock_mcp, test_workbook):
        """Test successful range copy."""
        result = copy_range_tool(TEST_WORKBOOK, TEST_SHEET, "A1", "B2", "C3")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Copied range" in result["message"]
        
    def test_copy_range_invalid_sheet(self, mock_mcp, test_workbook):
        """Test copying range from invalid sheet."""
        result = copy_range_tool(TEST_WORKBOOK, "InvalidSheet", "A1", "B2", "C3")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "Sheet" in result["message"]
        
    def test_copy_range_invalid_source(self, mock_mcp, test_workbook):
        """Test copying invalid source range."""
        result = copy_range_tool(TEST_WORKBOOK, TEST_SHEET, "InvalidRange", "B2", "C3")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "error" in result["message"].lower()
        
    def test_copy_range_invalid_target(self, mock_mcp, test_workbook):
        """Test copying to invalid target."""
        result = copy_range_tool(TEST_WORKBOOK, TEST_SHEET, "A1", "B2", "InvalidTarget")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "error" in result["message"].lower()

class TestMoveRangeTool:
    """Test cases for move_range_tool."""
    
    def test_move_range_success(self, mock_mcp, test_workbook):
        """Test successful range move."""
        result = move_range_tool(TEST_WORKBOOK, TEST_SHEET, TEST_RANGE, TEST_TARGET)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Moved range" in result["message"]
        
    def test_move_range_invalid_sheet(self, mock_mcp, test_workbook):
        """Test moving range from invalid sheet."""
        result = move_range_tool(TEST_WORKBOOK, "InvalidSheet", TEST_RANGE, TEST_TARGET)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "Sheet" in result["message"]
        
    def test_move_range_invalid_source(self, mock_mcp, test_workbook):
        """Test moving invalid source range."""
        result = move_range_tool(TEST_WORKBOOK, TEST_SHEET, "InvalidRange", TEST_TARGET)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "error" in result["message"].lower()
        
    def test_move_range_invalid_target(self, mock_mcp, test_workbook):
        """Test moving to invalid target."""
        result = move_range_tool(TEST_WORKBOOK, TEST_SHEET, TEST_RANGE, "InvalidTarget")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "error" in result["message"].lower()

class TestMergeRangeTool:
    """Test cases for merge_range_tool."""
    
    def test_merge_range_success(self, mock_mcp, test_workbook):
        """Test successful range merge."""
        result = merge_range_tool(TEST_WORKBOOK, TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Merged range" in result["message"]
        
    def test_merge_range_invalid_sheet(self, mock_mcp, test_workbook):
        """Test merging range in invalid sheet."""
        result = merge_range_tool(TEST_WORKBOOK, "InvalidSheet", "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "Sheet" in result["message"]
        
    def test_merge_range_invalid_range(self, mock_mcp, test_workbook):
        """Test merging invalid range."""
        result = merge_range_tool(TEST_WORKBOOK, TEST_SHEET, "InvalidRange", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "error" in result["message"].lower()

class TestUnmergeRangeTool:
    """Test cases for unmerge_range_tool."""
    
    def test_unmerge_range_success(self, mock_mcp, test_workbook):
        """Test successful range unmerge."""
        # First merge a range
        merge_range_tool(TEST_WORKBOOK, TEST_SHEET, "A1", "B2")
        
        # Then unmerge it
        result = unmerge_range_tool(TEST_WORKBOOK, TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Unmerged range" in result["message"]
        
    def test_unmerge_range_invalid_sheet(self, mock_mcp, test_workbook):
        """Test unmerging range in invalid sheet."""
        result = unmerge_range_tool(TEST_WORKBOOK, "InvalidSheet", "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "Sheet" in result["message"]
        
    def test_unmerge_range_invalid_range(self, mock_mcp, test_workbook):
        """Test unmerging invalid range."""
        result = unmerge_range_tool(TEST_WORKBOOK, TEST_SHEET, "InvalidRange", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "error" in result["message"].lower() 