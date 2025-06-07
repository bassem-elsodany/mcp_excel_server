"""
Test cases for Excel MCP Server excel tools.

This module contains test cases for all excel tools including:
- read_data_from_excel
- write_data_to_excel
- merge_cells
- unmerge_cells
- list_excel_files
"""

import os
import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment
import logging

from mcp.server.fastmcp import FastMCP
from mcp_excel_server.core.exceptions import ValidationError, DataError, SheetError
from mcp_excel_server.config.settings import settings

from mcp_excel_server.api.tools.excel import (
    read_data_from_excel,
    write_data_to_excel,
    merge_cells,
    unmerge_cells,
    list_excel_files
)

from mcp_excel_server.utils import get_logger
# Initialize logger
logger = logging.getLogger(__name__)

# Test data
TEST_WORKBOOK = "test_workbook.xlsx"
TEST_SHEET = "Sheet1"
TEST_RANGE = "A1:B2"
TEST_CELL = "A1"

@pytest.fixture
def mock_mcp():
    """Create a mock FastMCP server instance."""
    return Mock(spec=FastMCP)

@pytest.fixture
def test_workbook():
    """Create a test workbook with sample data."""
    logger.debug("Creating test workbook")
    wb = Workbook()
    logger.debug(f"Initial active sheet: {wb.active}")
    
    if wb.active is None:
        logger.debug("No active sheet, creating new sheet")
        ws = wb.create_sheet()
    else:
        logger.debug("Using existing active sheet")
        ws = wb.active
    
    logger.debug(f"Setting sheet title to {TEST_SHEET}")
    ws.title = TEST_SHEET
    
    # Add some test data
    logger.debug("Adding test data to worksheet")
    ws['A1'] = "Test1"
    ws['A2'] = "Test2"
    ws['B1'] = "Test3"
    ws['B2'] = "Test4"
    
    # Add some formatting
    logger.debug("Adding formatting to worksheet")
    ws['A1'].font = Font(bold=True)
    ws['B1'].number_format = "#,##0.00"
    
    # Save workbook
    logger.debug(f"Saving workbook to {TEST_WORKBOOK}")
    excel_path = os.path.join(settings.excel_mcp_folder, TEST_WORKBOOK)
    os.makedirs(settings.excel_mcp_folder, exist_ok=True)
    wb.save(excel_path)
    
    yield excel_path
    
    # Cleanup
    logger.debug("Cleaning up test workbook")
    if os.path.exists(excel_path):
        os.remove(excel_path)

class TestReadDataFromExcel:
    """Test cases for read_data_from_excel."""
    
    def test_read_data_success(self, mock_mcp, test_workbook):
        """Test successful data reading."""
        result = read_data_from_excel(TEST_WORKBOOK, TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "data" in result
        assert "Test1" in result["data"]
        assert "Test3" in result["data"]
        
    def test_read_data_preview(self, mock_mcp, test_workbook):
        """Test reading data in preview mode."""
        result = read_data_from_excel(TEST_WORKBOOK, TEST_SHEET, "A1", "B2", preview_only=True)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "data" in result
        assert "Test1" in result["data"]
        
    def test_read_data_invalid_workbook(self, mock_mcp):
        """Test reading from invalid workbook."""
        result = read_data_from_excel("nonexistent.xlsx", TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()
            
    def test_read_data_invalid_sheet(self, mock_mcp, test_workbook):
        """Test reading from invalid sheet."""
        result = read_data_from_excel(TEST_WORKBOOK, "NonexistentSheet", "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()

class TestWriteDataToExcel:
    """Test cases for write_data_to_excel."""
    
    def test_write_data_success(self, mock_mcp, test_workbook):
        """Test successful data writing."""
        test_data = [["New1", "New2"], ["New3", "New4"]]
        result = write_data_to_excel(TEST_WORKBOOK, TEST_SHEET, test_data, "C1")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "data written" in result["message"].lower()
        
    def test_write_data_append(self, mock_mcp, test_workbook):
        """Test appending data."""
        test_data = [["Append1", "Append2"]]
        result = write_data_to_excel(TEST_WORKBOOK, TEST_SHEET, test_data)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "data written" in result["message"].lower()
        
    def test_write_data_invalid_workbook(self, mock_mcp):
        """Test writing to invalid workbook."""
        result = write_data_to_excel("nonexistent.xlsx", TEST_SHEET, [["Test"]])
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()
            
    def test_write_data_invalid_sheet(self, mock_mcp, test_workbook):
        """Test writing to invalid sheet."""
        result = write_data_to_excel(TEST_WORKBOOK, "NonexistentSheet", [["Test"]])
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()

class TestMergeCells:
    """Test cases for merge_cells."""
    
    def test_merge_cells_success(self, mock_mcp, test_workbook):
        """Test successful cell merging."""
        result = merge_cells(TEST_WORKBOOK, TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "merged" in result["message"].lower()
        
    def test_merge_cells_invalid_workbook(self, mock_mcp):
        """Test merging cells in invalid workbook."""
        result = merge_cells("nonexistent.xlsx", TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()
            
    def test_merge_cells_invalid_sheet(self, mock_mcp, test_workbook):
        """Test merging cells in invalid sheet."""
        result = merge_cells(TEST_WORKBOOK, "NonexistentSheet", "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()

class TestUnmergeCells:
    """Test cases for unmerge_cells."""
    
    def test_unmerge_cells_success(self, mock_mcp, test_workbook):
        """Test successful cell unmerging."""
        # First merge the cells
        merge_cells(TEST_WORKBOOK, TEST_SHEET, "A1", "B2")
        
        result = unmerge_cells(TEST_WORKBOOK, TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "unmerged" in result["message"].lower()
        
    def test_unmerge_cells_invalid_workbook(self, mock_mcp):
        """Test unmerging cells in invalid workbook."""
        result = unmerge_cells("nonexistent.xlsx", TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()
            
    def test_unmerge_cells_invalid_sheet(self, mock_mcp, test_workbook):
        """Test unmerging cells in invalid sheet."""
        result = unmerge_cells(TEST_WORKBOOK, "NonexistentSheet", "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()

class TestListExcelFiles:
    """Test cases for list_excel_files."""
    
    def test_list_excel_files_success(self, mock_mcp, test_workbook):
        """Test successful file listing."""
        logger.debug("Testing successful file listing")
        result = list_excel_files()
        
        logger.debug(f"Result: {result}")
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "files" in result
        assert isinstance(result["files"], list)
        assert TEST_WORKBOOK in result["files"]
        
    def test_list_excel_files_empty(self, mock_mcp):
        """Test listing files when directory is empty."""
        logger.debug("Testing empty directory listing")
        with patch('os.listdir', return_value=[]):
            result = list_excel_files()
            
            logger.debug(f"Result: {result}")
            assert isinstance(result, dict)
            assert result["success"] is True
            assert len(result["files"]) == 0 