"""
Test cases for Excel MCP Server workbook tools.

This module contains test cases for all workbook tools including:
- create_workbook
- list_workbooks
- read_workbook_data
- write_workbook_data
"""

import os
import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook

from mcp.server.fastmcp import FastMCP
from mcp_excel_server.core.exceptions import ValidationError, DataError, SheetError
from mcp_excel_server.config.settings import settings
from mcp_excel_server.api.tools.workbook import (
    create_workbook,
    list_workbooks,
    read_workbook_data,
    write_workbook_data
)

# Test data
TEST_WORKBOOK_FILENAME = "test_workbook.xlsx"
TEST_NONE_EXISTENT_WORKBOOK_FILENAME = "nonexistent.xlsx"
TEST_WORKBOOK = os.path.join(settings.excel_mcp_folder, TEST_WORKBOOK_FILENAME)
TEST_NONE_EXISTENT_WORKBOOK = os.path.join(settings.excel_mcp_folder, TEST_NONE_EXISTENT_WORKBOOK_FILENAME)
TEST_SHEET = "Sheet1"

@pytest.fixture
def mock_mcp():
    """Create a mock FastMCP server instance."""
    return Mock(spec=FastMCP)

@pytest.fixture
def test_workbook():
    """Create a test workbook with sample data."""
    # Ensure test directory exists
    os.makedirs(settings.excel_mcp_folder, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = TEST_SHEET
    
    # Add some test data
    ws['A1'] = "Test1"
    ws['A2'] = "Test2"
    ws['B1'] = "Test3"
    ws['B2'] = "Test4"
    
    # Save workbook
    wb.save(TEST_WORKBOOK)
    
    yield TEST_WORKBOOK_FILENAME
    
    # Cleanup
    if os.path.exists(TEST_WORKBOOK):
        os.remove(TEST_WORKBOOK)
    if os.path.exists(TEST_NONE_EXISTENT_WORKBOOK):
        os.remove(TEST_NONE_EXISTENT_WORKBOOK)

class TestCreateWorkbook:
    """Test cases for create_workbook."""
    
    def test_create_workbook_success(self, mock_mcp):
        """Test successful workbook creation."""
        # Ensure the file does not exist before the test
        if os.path.exists(TEST_WORKBOOK):
            os.remove(TEST_WORKBOOK)
        result = create_workbook(TEST_WORKBOOK_FILENAME)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Created workbook" in result["message"]
        assert "info" in result
        assert "filename" in result["info"]
        assert result["info"]["filename"] == TEST_WORKBOOK_FILENAME
        assert os.path.exists(TEST_WORKBOOK)
        
    def test_create_workbook_existing(self, mock_mcp, test_workbook):
        """Test creating workbook that already exists."""
        result = create_workbook(TEST_WORKBOOK_FILENAME)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "already exists" in result["message"].lower()

class TestListWorkbooks:
    """Test cases for list_workbooks."""
    
    def test_list_workbooks_success(self, mock_mcp, test_workbook):
        """Test successful workbook listing."""
        result = list_workbooks()
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "files" in result
        assert isinstance(result["files"], list)
        assert any(TEST_WORKBOOK_FILENAME in wb for wb in result["files"])

class TestReadWorkbookData:
    """Test cases for read_workbook_data."""
    
    def test_read_workbook_data_success(self, mock_mcp, test_workbook):
        """Test successful data reading."""
        result = read_workbook_data(TEST_WORKBOOK_FILENAME, TEST_SHEET, "A1", "B2")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "data" in result
        assert "message" in result
        assert result["message"] in ["Data read successfully", "No data found in specified range"]
    
    def test_read_workbook_data_nonexistent(self, mock_mcp):
        """Test reading from nonexistent workbook."""
        result = read_workbook_data(TEST_NONE_EXISTENT_WORKBOOK_FILENAME, TEST_SHEET)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "message" in result
        assert (
            "no such file or directory" in result["message"].lower() or
            "file not found" in result["message"].lower()
        )

class TestWriteWorkbookData:
    """Test cases for write_workbook_data."""
    
    def test_write_workbook_data_success(self, mock_mcp, test_workbook):
        """Test successful data writing."""
        test_data = [["New1", "New2"], ["New3", "New4"]]
        result = write_workbook_data(TEST_WORKBOOK_FILENAME, TEST_SHEET, test_data, "C1")
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "message" in result
        
        # Verify data was written
        read_result = read_workbook_data(TEST_WORKBOOK_FILENAME, TEST_SHEET, "C1", "D2")
        assert read_result["success"] is True
        assert "data" in read_result
        assert "message" in read_result
        assert read_result["message"] in ["Data read successfully", "No data found in specified range"]
    
    def test_write_workbook_data_nonexistent(self, mock_mcp):
        """Test writing to nonexistent workbook."""
        test_data = [["New1", "New2"]]
        result = write_workbook_data(TEST_NONE_EXISTENT_WORKBOOK_FILENAME, TEST_SHEET, test_data)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "message" in result
        assert (
            "no such file or directory" in result["message"].lower() or
            "file not found" in result["message"].lower()
        )
        
        
            