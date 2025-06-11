"""
Test cases for Excel MCP Server worksheet tools.

This module contains test cases for all worksheet tools including:
- create_worksheet_tool
- delete_worksheet_tool
- rename_worksheet_tool
- copy_worksheet_tool
- move_worksheet_tool
"""

import os
import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook
from mcp_excel_server.config.settings import settings
from mcp.server.fastmcp import FastMCP
from mcp_excel_server.core.exceptions import ValidationError, SheetError

from mcp_excel_server.api.tools.worksheet import (
    create_worksheet,
    delete_worksheet,
    rename_worksheet,
    copy_worksheet,
    move_worksheet
)

# Test data
TEST_WORKBOOK_FILENAME = "test_workbook.xlsx"
TEST_WORKBOOK = os.path.join(settings.excel_mcp_folder, TEST_WORKBOOK_FILENAME)
TEST_INVALID_WORKBOOK = "/invalid/name.xlsx"
TEST_NONE_EXISTENT_WORKBOOK = "/nonexistent.xlsx"
TEST_SHEET = "Sheet1"
TEST_NEW_SHEET = "NewSheet"
TEST_COPY_SHEET = "CopySheet"

@pytest.fixture
def mock_mcp():
    """Create a mock FastMCP server instance."""
    return Mock(spec=FastMCP)

@pytest.fixture
def test_workbook():
    """Create a test workbook with some data."""
    # Ensure test directory exists
    os.makedirs(settings.excel_mcp_folder, exist_ok=True)
    
    # Create test workbook
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = TEST_SHEET
    
    # Add some test data
    ws['A1'] = "Test1"
    ws['A2'] = "Test2"
    ws['B1'] = "Test3"
    ws['B2'] = "Test4"
    
    # Save workbook
    wb.save(TEST_WORKBOOK)
    
    yield TEST_WORKBOOK_FILENAME
    
    # Cleanup
    if os.path.exists(TEST_WORKBOOK):
        os.remove(TEST_WORKBOOK)

class TestCreateWorksheetTool:
    """Test cases for create_worksheet_tool."""
    
    def test_create_worksheet_success(self, mock_mcp, test_workbook):
        """Test successful worksheet creation."""
        result = create_worksheet(test_workbook, TEST_NEW_SHEET)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Created worksheet" in result["message"]
        assert result["sheet_name"] == TEST_NEW_SHEET
        
    def test_create_worksheet_already_exists(self, mock_mcp, test_workbook):
        """Test creating worksheet that already exists."""
        # First create the worksheet
        create_worksheet(test_workbook, TEST_NEW_SHEET)
        
        # Try to create it again
        result = create_worksheet(test_workbook, TEST_NEW_SHEET)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "already exists" in result["message"].lower()
        

class TestDeleteWorksheetTool:
    """Test cases for delete_worksheet_tool."""
    
    def test_delete_worksheet_success(self, mock_mcp, test_workbook):
        """Test successful worksheet deletion."""
       
        # First create the worksheet
        create_worksheet(test_workbook, TEST_NEW_SHEET)

        result = delete_worksheet(test_workbook, TEST_SHEET)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Deleted worksheet" in result["message"]
        
    def test_delete_worksheet_not_found(self, mock_mcp, test_workbook):
        """Test deleting non-existent worksheet."""
         # First create the worksheet
        create_worksheet(test_workbook, TEST_NEW_SHEET)
        result = delete_worksheet(test_workbook, "NonExistentSheet")
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()
        

class TestRenameWorksheetTool:
    """Test cases for rename_worksheet_tool."""
    
    def test_rename_worksheet_success(self, mock_mcp, test_workbook):
        """Test successful worksheet rename."""
        result = rename_worksheet(test_workbook, TEST_SHEET, TEST_NEW_SHEET)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Renamed worksheet" in result["message"]
        assert result["new_name"] == TEST_NEW_SHEET
        
    def test_rename_worksheet_not_found(self, mock_mcp, test_workbook):
        """Test renaming non-existent worksheet."""
        result = rename_worksheet(test_workbook, "NonExistentSheet", TEST_NEW_SHEET)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()
        

class TestCopyWorksheetTool:
    """Test cases for copy_worksheet_tool."""
    
    def test_copy_worksheet_success(self, mock_mcp, test_workbook):
        """Test successful worksheet copy."""
        result = copy_worksheet(test_workbook, TEST_SHEET, TEST_COPY_SHEET)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Copied worksheet" in result["message"]
        assert result["new_name"] == TEST_COPY_SHEET
        
    def test_copy_worksheet_not_found(self, mock_mcp, test_workbook):
        """Test copying non-existent worksheet."""
        result = copy_worksheet(test_workbook, "NonExistentSheet", TEST_COPY_SHEET)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()
        

class TestMoveWorksheetTool:
    """Test cases for move_worksheet_tool."""
    
    def test_move_worksheet_success(self, mock_mcp, test_workbook):
        """Test successful worksheet move."""
         # First create the worksheet
        create_worksheet(test_workbook, TEST_NEW_SHEET)
        result = move_worksheet(test_workbook, TEST_SHEET, 1)
        
        assert isinstance(result, dict)
        assert result["success"] is True
        assert "Moved worksheet" in result["message"]
        assert result["sheet_name"] == TEST_SHEET
        
    def test_move_worksheet_not_found(self, mock_mcp, test_workbook):
        """Test moving non-existent worksheet."""
        result = move_worksheet(test_workbook, "NonExistentSheet", 1)
        
        assert isinstance(result, dict)
        assert result["success"] is False
        assert "not found" in result["message"].lower()