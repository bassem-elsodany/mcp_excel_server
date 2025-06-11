"""
Range management module for Excel MCP Server.
This module provides functions for managing ranges in Excel workbooks.
"""

from typing import Dict, Any, Optional, Union
from pathlib import Path
from openpyxl import Workbook

from mcp_excel_server.core.workbook import get_workbook, save_workbook
from mcp_excel_server.core.exceptions import RangeError
from mcp_excel_server.utils import get_logger
from mcp_excel_server.utils.logger import audit_event
from mcp_excel_server.config.settings import settings
from mcp_excel_server.core.sheet import (
    copy_range_operation,
    delete_range_operation,
    merge_range as sheet_merge_range,
    unmerge_range as sheet_unmerge_range
)

logger = get_logger(__name__)

def get_full_path(filename: str) -> Path:
    """Get the full path for a workbook file."""
    return Path(settings.excel_mcp_folder) / filename

def delete_range(
    filename: str,
    sheet_name: str,
    range_str: str
) -> Dict[str, Any]:
    """Delete a range of cells in an Excel workbook.
    
    Args:
        filename: Name of the Excel file
        sheet_name: Name of the sheet containing the range
        range_str: Range to delete (e.g. 'A1:B10')
        
    Returns:
        Dict containing operation result
    """
    try:
        path = get_full_path(filename)
        result = delete_range_operation(str(path), sheet_name, range_str)
        audit_event("delete_range", {
            "file": str(path),
            "sheet": sheet_name,
            "range": range_str
        })
        return result
    except Exception as e:
        logger.error(f"Failed to delete range: {e}")
        raise RangeError(str(e))

def copy_range(
    filename: str,
    sheet_name: str,
    source_range: str,
    target_range: str
) -> Dict[str, Any]:
    """Copy a range of cells in an Excel workbook.
    
    Args:
        filename: Name of the Excel file
        sheet_name: Name of the sheet containing the range
        source_range: Range to copy (e.g. 'A1:B10')
        target_range: Range to copy to (e.g. 'D1:E10')
        
    Returns:
        Dict containing operation result
    """
    try:
        path = get_full_path(filename)
        result = copy_range_operation(
            str(path),
            sheet_name,
            source_range,
            source_range,  # Using same range for start/end
            target_range
        )
        audit_event("copy_range", {
            "file": str(path),
            "sheet": sheet_name,
            "source_range": source_range,
            "target_range": target_range
        })
        return result
    except Exception as e:
        logger.error(f"Failed to copy range: {e}")
        raise RangeError(str(e))

def move_range(
    filename: str,
    sheet_name: str,
    source_range: str,
    target_range: str
) -> Dict[str, Any]:
    """Move a range of cells in an Excel workbook.
    
    Args:
        filename: Name of the Excel file
        sheet_name: Name of the sheet containing the range
        source_range: Range to move (e.g. 'A1:B10')
        target_range: Range to move to (e.g. 'D1:E10')
        
    Returns:
        Dict containing operation result
    """
    try:
        path = get_full_path(filename)
        # First copy to target
        result = copy_range_operation(
            str(path),
            sheet_name,
            source_range,
            source_range,
            target_range
        )
        # Then delete source
        delete_range_operation(str(path), sheet_name, source_range)
        
        audit_event("move_range", {
            "file": str(path),
            "sheet": sheet_name,
            "source_range": source_range,
            "target_range": target_range
        })
        return {"message": f"Range moved from {source_range} to {target_range}"}
    except Exception as e:
        logger.error(f"Failed to move range: {e}")
        raise RangeError(str(e))

def merge_range(
    filename: str,
    sheet_name: str,
    range_str: str
) -> Dict[str, Any]:
    """Merge a range of cells in an Excel workbook.
    
    Args:
        filename: Name of the Excel file
        sheet_name: Name of the sheet containing the range
        range_str: Range to merge (e.g. 'A1:B10')
        
    Returns:
        Dict containing operation result
    """
    try:
        path = get_full_path(filename)
        if ':' not in range_str:
            raise RangeError("Range must include start and end cells (e.g. 'A1:B10')")
            
        start_cell, end_cell = range_str.split(':')
        result = sheet_merge_range(str(path), sheet_name, start_cell, end_cell)
        
        audit_event("merge_range", {
            "file": str(path),
            "sheet": sheet_name,
            "range": range_str
        })
        return result
    except Exception as e:
        logger.error(f"Failed to merge range: {e}")
        raise RangeError(str(e))

def unmerge_range(
    filename: str,
    sheet_name: str,
    range_str: str
) -> Dict[str, Any]:
    """Unmerge a range of cells in an Excel workbook.
    
    Args:
        filename: Name of the Excel file
        sheet_name: Name of the sheet containing the range
        range_str: Range to unmerge (e.g. 'A1:B10')
        
    Returns:
        Dict containing operation result
    """
    try:
        path = get_full_path(filename)
        if ':' not in range_str:
            raise RangeError("Range must include start and end cells (e.g. 'A1:B10')")
            
        start_cell, end_cell = range_str.split(':')
        result = sheet_unmerge_range(str(path), sheet_name, start_cell, end_cell)
        
        audit_event("unmerge_range", {
            "file": str(path),
            "sheet": sheet_name,
            "range": range_str
        })
        return result
    except Exception as e:
        logger.error(f"Failed to unmerge range: {e}")
        raise RangeError(str(e))

def validate_range(
    filename: str,
    sheet_name: str,
    range_str: str
) -> Dict[str, Any]:
    """Validate a range in an Excel workbook.
    
    Args:
        filename: Name of the Excel file
        sheet_name: Name of the sheet containing the range
        range_str: Range to validate (e.g. 'A1:B10')
        
    Returns:
        Dict with range information including:
        - start_cell: str first cell in range
        - end_cell: str last cell in range
        - num_rows: int number of rows
        - num_cols: int number of columns
    """
    try:
        path = get_full_path(filename)
        wb = get_workbook(str(path))
        
        if sheet_name not in wb.sheetnames:
            raise RangeError(f"Sheet '{sheet_name}' not found")
            
        if ':' not in range_str:
            raise RangeError("Range must include start and end cells (e.g. 'A1:B10')")
            
        start_cell, end_cell = range_str.split(':')
        ws = wb[sheet_name]
        
        # Validate cells exist
        if not start_cell or not end_cell:
            raise RangeError("Invalid range format")
            
        # Get cell coordinates
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        start_col = int(''.join(filter(str.isalpha, start_cell)))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        end_col = int(''.join(filter(str.isalpha, end_cell)))
        
        # Validate against sheet dimensions
        if start_row > ws.max_row or end_row > ws.max_row:
            raise RangeError(f"Row out of bounds (1-{ws.max_row})")
        if start_col > ws.max_column or end_col > ws.max_column:
            raise RangeError(f"Column out of bounds (1-{ws.max_column})")
            
        return {
            "start_cell": start_cell,
            "end_cell": end_cell,
            "num_rows": end_row - start_row + 1,
            "num_cols": end_col - start_col + 1,
            "is_valid": True
        }
    except Exception as e:
        logger.error(f"Failed to validate range: {e}")
        raise RangeError(str(e)) 