"""
Sheet operations for Excel MCP Server.
"""

from typing import Any, Optional
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Border, PatternFill, Side, Alignment

from mcp_excel_server.core.cell_utils import parse_cell_range
from mcp_excel_server.core.exceptions import SheetError, ValidationError
from mcp_excel_server.utils import get_logger
from mcp_excel_server.utils.logger import audit_event
from mcp_excel_server.config.settings import settings

logger = get_logger(__name__)

def get_full_path(filename: str) -> Path:
    """Get the full path for a workbook file."""
    return Path(settings.excel_mcp_folder) / filename

def copy_sheet(filepath: str, source_sheet: str, target_sheet: str) -> dict[str, Any]:
    """Copy a worksheet within the same workbook."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if source_sheet not in wb.sheetnames:
            raise SheetError(f"Source sheet '{source_sheet}' not found")
            
        if target_sheet in wb.sheetnames:
            raise SheetError(f"Target sheet '{target_sheet}' already exists")
            
        source = wb[source_sheet]
        target = wb.copy_worksheet(source)
        target.title = target_sheet
        
        wb.save(str(path))
        return {"message": f"Sheet '{source_sheet}' copied to '{target_sheet}'"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to copy sheet: {e}")
        raise SheetError(str(e))

def delete_sheet(filepath: str, sheet_name: str) -> dict[str, Any]:
    """Delete a worksheet from the workbook."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        if len(wb.sheetnames) == 1:
            raise SheetError("Cannot delete the only sheet in workbook")
            
        del wb[sheet_name]
        wb.save(str(path))
        return {"message": f"Sheet '{sheet_name}' deleted"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete sheet: {e}")
        raise SheetError(str(e))

def rename_sheet(filepath: str, old_name: str, new_name: str) -> dict[str, Any]:
    """Rename a worksheet."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if old_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{old_name}' not found")
            
        if new_name in wb.sheetnames:
            raise SheetError(f"Sheet '{new_name}' already exists")
            
        sheet = wb[old_name]
        sheet.title = new_name
        wb.save(str(path))
        return {"message": f"Sheet renamed from '{old_name}' to '{new_name}'"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to rename sheet: {e}")
        raise SheetError(str(e))

def format_range_string(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    """Format range string from row and column indices."""
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

def copy_range(
    source_ws: Worksheet,
    target_ws: Worksheet,
    source_range: str,
    target_start: str | None = None,
) -> None:
    """Copy range from source worksheet to target worksheet."""
    # Parse source range
    if ':' in source_range:
        source_start, source_end = source_range.split(':')
    else:
        source_start = source_range
        source_end = None
        
    src_start_row, src_start_col, src_end_row, src_end_col = parse_cell_range(
        source_start, source_end
    )

    if src_end_row is None:
        src_end_row = src_start_row
    if src_end_col is None:
        src_end_col = src_start_col

    if target_start is None:
        target_start = source_start

    tgt_start_row, tgt_start_col, _, _ = parse_cell_range(target_start)

    for i, row in enumerate(range(src_start_row, src_end_row + 1)):
        for j, col in enumerate(range(src_start_col, src_end_col + 1)):
            source_cell = source_ws.cell(row=row, column=col)
            target_cell = target_ws.cell(row=tgt_start_row + i, column=tgt_start_col + j)

            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
                target_cell.border = Border(
                    left=copy(source_cell.border.left),
                    right=copy(source_cell.border.right),
                    top=copy(source_cell.border.top),
                    bottom=copy(source_cell.border.bottom)
                )
                target_cell.fill = PatternFill(
                    patternType=source_cell.fill.patternType,
                    fgColor=source_cell.fill.fgColor,
                    bgColor=source_cell.fill.bgColor
                )
                target_cell.number_format = source_cell.number_format
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )

def delete_range(worksheet: Worksheet, start_cell: str, end_cell: str | None = None) -> None:
    """Delete contents and formatting of a range."""
    start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

    if end_row is None:
        end_row = start_row
    if end_col is None:
        end_col = start_col

    for i in range(start_row, end_row + 1):
        for j in range(start_col, end_col + 1):
            cell = worksheet.cell(row=i, column=j)
            cell.value = None
            cell.font = Font()
            cell.border = Border()
            cell.fill = PatternFill()
            cell.number_format = "General"
            cell.alignment = copy(Alignment())

def merge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> dict[str, Any]:
    """Merge a range of cells."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

        if end_row is None or end_col is None:
            raise SheetError("Both start and end cells must be specified for merging")

        range_string = format_range_string(start_row, start_col, end_row, end_col)
        worksheet = wb[sheet_name]
        worksheet.merge_cells(range_string)
        wb.save(str(path))
        result = {"success": True, "message": f"Merged range {range_string} in {sheet_name}"}
        audit_event("merge", {
            "file": str(path),
            "sheet": sheet_name,
            "start_cell": start_cell,
            "end_cell": end_cell,
            "range": range_string
        })
        return result
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to merge range: {e}")
        raise SheetError(str(e))

def unmerge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> dict[str, Any]:
    """Unmerge a range of cells."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        
        if end_row is None or end_col is None:
            raise SheetError("Both start and end cells must be specified for unmerging")

        range_string = format_range_string(start_row, start_col, end_row, end_col)
        
        # Check if range is actually merged
        merged_ranges = worksheet.merged_cells.ranges
        target_range = range_string.upper()
        
        if not any(str(merged_range).upper() == target_range for merged_range in merged_ranges):
            raise SheetError(f"Range '{range_string}' is not merged")
            
        worksheet.unmerge_cells(range_string)
        wb.save(str(path))
        result = {"success": True, "message": f"Unmerged range {range_string} in {sheet_name}"}
        audit_event("unmerge", {
            "file": str(path),
            "sheet": sheet_name,
            "start_cell": start_cell,
            "end_cell": end_cell,
            "range": range_string
        })
        return result
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to unmerge range: {e}")
        raise SheetError(str(e))

def copy_range_operation(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> dict:
    """Copy a range of cells to another location."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found")
            raise ValidationError(f"Sheet '{sheet_name}' not found")

        source_ws = wb[sheet_name]
        target_ws = wb[target_sheet] if target_sheet else source_ws

        # Parse source range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(source_start, source_end)
            if end_row is None:
                end_row = start_row
            if end_col is None:
                end_col = start_col
        except ValueError as e:
            logger.error(f"Invalid source range: {e}")
            raise ValidationError(f"Invalid source range: {str(e)}")

        # Parse target starting point
        try:
            target_row = int(''.join(filter(str.isdigit, target_start)))
            target_col = column_index_from_string(''.join(filter(str.isalpha, target_start)))
        except ValueError as e:
            logger.error(f"Invalid target cell: {e}")
            raise ValidationError(f"Invalid target cell: {str(e)}")

        # Copy the range
        row_offset = target_row - start_row
        col_offset = target_col - start_col

        for i in range(start_row, end_row + 1):
            for j in range(start_col, end_col + 1):
                source_cell = source_ws.cell(row=i, column=j)
                target_cell = target_ws.cell(row=i + row_offset, column=j + col_offset)
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        color=source_cell.font.color
                    )
                    target_cell.border = Border(
                        left=copy(source_cell.border.left),
                        right=copy(source_cell.border.right),
                        top=copy(source_cell.border.top),
                        bottom=copy(source_cell.border.bottom)
                    )
                    target_cell.fill = PatternFill(
                        patternType=source_cell.fill.patternType,
                        fgColor=source_cell.fill.fgColor,
                        bgColor=source_cell.fill.bgColor
                    )
                    target_cell.number_format = source_cell.number_format
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        wrap_text=source_cell.alignment.wrap_text
                    )

        wb.save(str(path))
        return {"message": f"Range copied successfully"}

    except (ValidationError, SheetError):
        raise
    except Exception as e:
        logger.error(f"Failed to copy range: {e}")
        raise SheetError(f"Failed to copy range: {str(e)}")

def delete_range_operation(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
    shift_direction: str = "up"
) -> dict[str, Any]:
    """Delete a range of cells and shift remaining cells."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        # Validate range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            if end_row and end_row > worksheet.max_row:
                raise SheetError(f"End row {end_row} out of bounds (1-{worksheet.max_row})")
            if end_col and end_col > worksheet.max_column:
                raise SheetError(f"End column {end_col} out of bounds (1-{worksheet.max_column})")
        except ValueError as e:
            raise SheetError(f"Invalid range: {str(e)}")
            
        # Validate shift direction
        if shift_direction not in ["up", "left"]:
            raise ValidationError(f"Invalid shift direction: {shift_direction}. Must be 'up' or 'left'")
            
        range_string = format_range_string(
            start_row, start_col,
            end_row or start_row,
            end_col or start_col
        )
        
        # Delete range contents
        delete_range(worksheet, start_cell, end_cell)
        
        # Shift cells if needed
        if shift_direction == "up":
            worksheet.delete_rows(start_row, (end_row or start_row) - start_row + 1)
        elif shift_direction == "left":
            worksheet.delete_cols(start_col, (end_col or start_col) - start_col + 1)
            
        wb.save(str(path))
        
        return {"message": f"Range {range_string} deleted successfully"}
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete range: {e}")
        raise SheetError(str(e))

def create_sheet(filepath: str, sheet_name: str) -> dict[str, Any]:
    """Create a new worksheet in the workbook if it doesn't exist."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if sheet_name in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' already exists")
        wb.create_sheet(sheet_name)
        wb.save(str(path))
        return {"message": f"Sheet '{sheet_name}' created successfully"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create sheet: {e}")
        raise SheetError(str(e))

def move_sheet(filepath: str, sheet_name: str, index: int) -> dict[str, Any]:
    """Move a worksheet to a new position in the workbook.
    
    Args:
        filepath: Path to the Excel file
        sheet_name: Name of the sheet to move
        index: New position for the sheet (0-based)
        
    Returns:
        Dict containing:
        - message: str describing the result
    """
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        if index < 0 or index >= len(wb.sheetnames):
            raise SheetError(f"Invalid index {index}. Must be between 0 and {len(wb.sheetnames)-1}")
            
        sheet = wb[sheet_name]
        wb.move_sheet(sheet_name, offset=index - wb.index(sheet))
        wb.save(str(path))
        return {"message": f"Moved sheet '{sheet_name}' to position {index}"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to move sheet: {e}")
        raise SheetError(str(e))

def get_sheet(filepath: str, sheet_name: str) -> dict[str, Any]:
    """Get a worksheet by name from the workbook.
    
    Args:
        filepath: Path to the Excel file
        sheet_name: Name of the sheet to get
        
    Returns:
        Dict containing:
        - message: str describing the result
        - sheet: Worksheet object
    """
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        sheet = wb[sheet_name]
        return {
            "message": f"Retrieved sheet '{sheet_name}'",
            "sheet": sheet
        }
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get sheet: {e}")
        raise SheetError(str(e))

def list_sheets(filepath: str) -> list[Any]:
    """List all worksheets in the workbook.
    
    Args:
        filepath: Path to the Excel file
    Returns:
        List of worksheet objects
    """
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))
        return [wb[sheet_name] for sheet_name in wb.sheetnames]
    except Exception as e:
        logger.error(f"Failed to list sheets: {e}")
        raise SheetError(str(e))
