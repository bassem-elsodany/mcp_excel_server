"""
Workbook utilities for Excel MCP Server.
"""

from pathlib import Path
from typing import Any, Dict
import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from mcp_excel_server.core.exceptions import WorkbookError
from mcp_excel_server.utils import get_logger
from mcp_excel_server.config.settings import settings

logger = get_logger(__name__)

# Dictionary to store active workbooks
active_workbooks: Dict[str, Workbook] = {}

def get_full_path(filename: str) -> Path:
    """Get the full path for a workbook file."""
    return Path(settings.excel_mcp_folder) / filename

def create_workbook(filename: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
    """Create a new Excel workbook.

    Args:
        filename: Name of the Excel file to create
        sheet_name: Name of the initial worksheet

    Returns:
        Dict containing:
            success (bool): True if the workbook was created successfully
            message (str): A message describing the result
            filename (str): Name of the created workbook (if successful)
    """
    try:
        path = Path(settings.excel_mcp_folder) / filename
        if path.exists():
            return {
                "success": False,
                "message": f"Workbook already exists: {filename}",
                "filename": None
            }
        
        wb = Workbook()
        # Rename default sheet
        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            sheet.title = sheet_name
        else:
            wb.create_sheet(sheet_name)

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return {
            "success": True,
            "message": f"Created workbook '{filename}'",
            "filename": filename
        }
    except Exception as e:
        return {
            "success": False,
            "message": str(e),
            "filename": None
        }

def get_or_create_workbook(filepath: str) -> Workbook:
    """Get existing workbook or create new one if it doesn't exist"""
    try:
        path = get_full_path(filepath)
        return load_workbook(str(path))
    except FileNotFoundError:
        return create_workbook(filepath)["workbook"]

def create_sheet(filepath: str, sheet_name: str) -> dict:
    """Create a new worksheet in the workbook if it doesn't exist."""
    try:
        path = get_full_path(filepath)
        wb = load_workbook(str(path))

        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            raise WorkbookError(f"Sheet {sheet_name} already exists")

        # Create new sheet
        wb.create_sheet(sheet_name)
        wb.save(str(path))
        wb.close()
        return {"message": f"Sheet {sheet_name} created successfully"}
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create sheet: {e}")
        raise WorkbookError(str(e))

def get_workbook_info(filepath: str, include_ranges: bool = False) -> dict[str, Any]:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        path = get_full_path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")
            
        wb = load_workbook(str(path), read_only=False)
        
        info = {
            "filename": path.name,
            "sheets": wb.sheetnames,
            "size": path.stat().st_size,
            "modified": path.stat().st_mtime
        }
        
        if include_ranges:
            # Add used ranges for each sheet
            ranges = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 0 and ws.max_column > 0:
                    ranges[sheet_name] = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            info["used_ranges"] = ranges
            
        wb.close()
        return info
        
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get workbook info: {e}")
        raise WorkbookError(str(e))

def get_workbook(filepath: str) -> Workbook:
    """Get an existing workbook or raise an error if it doesn't exist."""
    try:
        path = get_full_path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")
        return load_workbook(str(path))
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get workbook: {e}")
        raise WorkbookError(str(e))

def save_workbook(wb: Workbook, filepath: str) -> None:
    """Save the given workbook to the specified filepath."""
    try:
        path = get_full_path(filepath)
        wb.save(str(path))
    except Exception as e:
        logger.error(f"Failed to save workbook: {e}")
        raise WorkbookError(f"Failed to save workbook: {e!s}")

def list_workbooks() -> Dict[str, Any]:
    """
    Lists all Excel files (.xlsx and .xlsm) in the configured directory (EXCEL_MCP_FOLDER).
    The directory will be created if it doesn't exist.

    Returns:
        dict: {
            "success": bool,  # True if the operation succeeded
            "files": list[str],  # List of Excel filenames (e.g., ["file1.xlsx", "file2.xlsm"])
            "message": str (optional)  # Error message if success is False
        }

    Example:
        list_workbooks()
        # -> { "success": True, "files": ["report.xlsx", "data.xlsm"] }
        # -> { "success": False, "message": "Permission denied", "files": [] }
    """
    try:
        logger.debug(f"Listing Excel files in {settings.excel_mcp_folder}")
        # Ensure directory exists
        os.makedirs(settings.excel_mcp_folder, exist_ok=True)
        
        files = []
        for f in os.listdir(settings.excel_mcp_folder):
            if f.endswith('.xlsx') or f.endswith('.xlsm'):
                files.append(f)
        logger.debug(f"Found files: {files}")
        return {"success": True, "files": files}
    except Exception as e:
        logger.error(f"Error listing Excel files: {e}")
        return {"success": False, "message": str(e), "files": []}
